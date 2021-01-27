import os
from openpyxl import load_workbook
workbook = load_workbook(filename="spreadsheet.xlsx")
sheet = workbook.active
from openpyxl import Workbook

wb = Workbook ()

ws = wb.active

ws.append (["Name", "USN", "Marks 1", "Marks 2","Marks 3"])

for i in range(2):
    if(int(sheet.cell(row=i+2, column=3).value)<60 or int(sheet.cell(row=i+2, column=4).value)<60 or int(sheet.cell(row=i+2, column=5).value)<60):
        print((sheet.cell(row=i+2, column=3).value))

        print("Yes")
        ws.append ([sheet.cell(row=i+2, column=1).value,sheet.cell(row=i+2, column=2).value,sheet.cell(row=i+2, column=3).value,sheet.cell(row=i+2, column=4).value,sheet.cell(row=i+2, column=5).value])
    else:
        print("No")    
wb.save ("spreadsheet1.xlsx")

os. system("spreadsheet1.xlsx")
